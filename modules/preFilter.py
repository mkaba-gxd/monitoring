import os
import sys
import pandas as pd
import numpy as np
from itertools import product
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from pathlib import Path
from .func import *

file_oncoKB = '/data1/GxD_eWES/reference/anno_db/oncoKB/v4.23/variant_table.tsv'

hidden_columns_ewes_1 = ['ID','QUAL','FILTER','GT','REF_AD','ALT_AD','DP','Allele','IMPACT','Gene','Feature_type','BIOTYPE','EXON','INTRON','cDNA_position','CDS_position','Protein_position','Amino_acids','Codons','Existing_variation','DISTANCE','STRAND','FLAGS','VARIANT_CLASS','SYMBOL_SOURCE','HGNC_ID','CANONICAL','MANE_PLUS_CLINICAL','TSL','APPRIS','CCDS','ENSP','SWISSPROT','TREMBL','UNIPARC','UNIPROT_ISOFORM','SOURCE','GENE_PHENO','SIFT','PolyPhen','DOMAINS','miRNA','HGVS_OFFSET','HGVSg','Clinvar','Clinvar_ALLELEID','Clinvar_CLNDN','Clinvar_CLNHGVS','Clinvar_CLNVC','Clinvar_GENEINFO','Clinvar_MC','FILTER_DEPTH','FILTER_ALT_DEPTH','FILTER_VAF','FILTER_AF_gnomADg','FILTER_AF_gnomADe','FILTER_AF_tommo','FILTER_NONCODING','FILTER_SPLICING','FILTER_SYNONYMOUS','FILTER_BLACKLIST']
hidden_columns_ewes_2 = ['ID','QUAL','FILTER','GT','REF_AD','ALT_AD','DP','Allele','IMPACT','Gene','Feature_type','BIOTYPE','EXON','INTRON','cDNA_position','CDS_position','Protein_position','Amino_acids','Codons','Existing_variation','DISTANCE','STRAND','FLAGS','VARIANT_CLASS','SYMBOL_SOURCE','HGNC_ID','CANONICAL','MANE_PLUS_CLINICAL','TSL','APPRIS','CCDS','ENSP','SWISSPROT','TREMBL','UNIPARC','UNIPROT_ISOFORM','SOURCE','GENE_PHENO','SIFT','PolyPhen','DOMAINS','miRNA','HGVS_OFFSET','HGVSg','Clinvar','Clinvar_ALLELEID','Clinvar_CLNDN','Clinvar_CLNHGVS','Clinvar_CLNVC','Clinvar_GENEINFO','Clinvar_MC','FILTER_BLACKLIST']
use_columns_ewes = ['CHROM','POS','ID','REF','ALT','QUAL','FILTER','GT','REF_AD','ALT_AD','AF','DP','Allele','Consequence','IMPACT','SYMBOL','Gene','Feature_type','Feature','BIOTYPE','EXON','INTRON','HGVSc','HGVSp','cDNA_position','CDS_position','Protein_position','Amino_acids','Codons','Existing_variation','DISTANCE','STRAND','FLAGS','VARIANT_CLASS','SYMBOL_SOURCE','HGNC_ID','CANONICAL','MANE_SELECT','MANE_PLUS_CLINICAL','TSL','APPRIS','CCDS','ENSP','SWISSPROT','TREMBL','UNIPARC','UNIPROT_ISOFORM','SOURCE','GENE_PHENO','SIFT','PolyPhen','DOMAINS','miRNA','HGVS_OFFSET','HGVSg','Clinvar','Clinvar_CLNSIG','Clinvar_ALLELEID','Clinvar_CLNDN','Clinvar_CLNHGVS','Clinvar_CLNVC','Clinvar_GENEINFO','Clinvar_MC','ONCOKB_ONCOGENICITY','ONCOKB_VARIANT_KEY','FILTER_DEPTH','FILTER_ALT_DEPTH','FILTER_VAF','FILTER_AF_gnomADg','FILTER_AF_gnomADe','FILTER_AF_tommo','FILTER_NONCODING','FILTER_SPLICING','FILTER_SYNONYMOUS','FILTER_BLACKLIST']

hidden_columns_ewes_3 = ['CN','C.flagged','seg.mean','seg.id','number.targets','gene.mean','gene.min','gene.max','focal','breakpoints','num.snps','M','M.flagged','loh']
use_columns_ewes_2 = ['Gene_name','CHROM','START','END','CN','C.flagged','seg.mean','seg.id','number.targets','gene.mean','gene.min','gene.max','focal','breakpoints','TYPE','num.snps','M','M.flagged','loh','ONCOKB_VARIANT_CNV','gene.mean.CN','FILTER_CN','FILTER_ONCOKB','FILTER_GENES']

use_columns_wts_1 = ['Out-of-Frame','OncoKB','cancer-related','gene1','gene2','chr1','breakpoint_1','chr2','breakpoint_2','max_split_cnt','max_span_cnt','sample_type','disease','tools','inferred_fusion_type','samples','cancer_db_hits','fusion_IDs']

def expand_breakpoints(df):

    expanded_rows = []

    for _, row in df.iterrows():

        bp1_raw = str(row['breakpoint_1'])
        bp2_raw = str(row['breakpoint_2'])

        bp1_values = bp1_raw.split('|') if '|' in bp1_raw else [bp1_raw]
        bp2_values = bp2_raw.split('|') if '|' in bp2_raw else [bp2_raw]

        for bp1, bp2 in product(bp1_values, bp2_values):
            new_row = row.copy()
            new_row['breakpoint_1'] = bp1
            new_row['breakpoint_2'] = bp2
            expanded_rows.append(new_row)

    return pd.DataFrame(expanded_rows)

def cull_columns(file, sheet_name, hidden_columns, data) :
    wb = load_workbook(file)
    ws = wb[sheet_name]
    for col_name in hidden_columns:
        if col_name in data.columns:
            col_idx = data.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True
    wb.save(file)

def add_info(file, sheet_name, info):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    df = pd.DataFrame(ws.values)
    start_row = df.shape[0] + 1

    for val in info :
        start_row = start_row + 1
        if type(val) is str:
            ws["A"+str(start_row)] = val
        elif type(val) is list :
            for c in range(len(val)):
                col_letter = get_column_letter(c+1)
                ws[ col_letter+str(start_row) ] = val[c]

    wb.save(file)

def coloring(file, sheet_name, column, val, color="FFFF00"):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df.drop(df.index[0]).reset_index(drop=True)

    if type(column) is str and type(val) is str:
        if not column in df.columns : return
        col_idx = df.index[df[column] == val].tolist()

    elif type(column) is list and type(val) is str:
        if not set(column).issubset(df.columns) : return
        cond = pd.Series([True] * len(df))
        for c in column:
            cond &= (df[c] == val)
        col_idx = df.index[cond].tolist()

    elif type(column) is list and type(val) is list:
        if not set(column).issubset(df.columns) : return
        cond = pd.Series([True] * len(df))
        for c, v in zip(column, val):
            cond &= (df[c] == v)
        col_idx = df.index[cond].tolist()

    else :
        return

    if len(col_idx) == 0 : return

    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row_idx in col_idx:
        excel_row = row_idx + 2 
        for c in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=c).fill = fill

    wb.save(file)

def add_filt(file, sheet_name) :
    wb = load_workbook(file)
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)
    ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

    wb.save(file)


def auto_filtering(file, sheet_name, column, val):
    wb = load_workbook(file)
    ws = wb[sheet_name]

    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])

    if not column in df.columns :
        print('No applicable columns: '+column)
        return

    for idx, value in enumerate(df[column], start=2):
        if value != val:
            ws.row_dimensions[idx].hidden = True

    wb.save(file)

def pic_value(file, column):
    try :
        df = pd.read_csv(file, sep="\t")
        if not column in df.columns :
            return '-'
        else :
            return df[column][0]
    except Exception as e:
        return '-'

def remove_files(FILES) :
    for file in FILES:
        if os.path.isfile(file):
            os.remove(file)
    
def run_preFilter(args):

    flowcellid = args.flowcellid
    directory = args.directory
    project_type = args.project_type
    outdir = args.outdir
    inclusion = [x.strip() for x in args.inclusion.split(',') if not x.strip() == '']
    exclusion = [x.strip() for x in args.exclusion.split(',') if not x.strip() == '']

    if len(inclusion) > 0 and len(exclusion) > 0 :
        init('ERROR: Inclusion and exclusion cannot be specified simultaneously.')

    var_tbl = pd.read_csv(file_oncoKB, sep="\t")
    var_tbl = var_tbl[var_tbl['VARIANT_GROUP_ID']=='SNV']
    var_tbl = var_tbl[['HUGO_SYMBOL','GRCH38_REFSEQ','GRCH38_ISOFORM']].drop_duplicates()
    var_tbl.columns = ['SYMBOL','ONCOKB_REFSEQ','ONCOKB_GENCODE']

    df_info = getinfo(SelectData(flowcellid))
    if df_info.shape[0] == 0 : init("No matching data found.")

    if len(inclusion) > 0:
        print ("inclusion sample:" + "\n".join(inclusion))
        df_info = df_info[ df_info['SAMPLE_ID'].isin(inclusion)]
        if df_info.shape[0] == 0 : init("No corresponding sample IDs.")

    if len(exclusion) > 0:
        print ("exclusion sample:" + ",".join(exclusion))
        df_info = df_info[ ~df_info['SAMPLE_ID'].isin(exclusion)]
        if df_info.shape[0] == 0 : init("No corresponding sample IDs.")

    df_info['PRJ_TYPE'] = df_info['PRJ_TYPE'].str.replace('EWES',"eWES")
    if project_type == "both" :
        df_info = df_info[ df_info['PRJ_TYPE'].isin(['eWES','WTS']) ]
    else :
        df_info = df_info[ df_info['PRJ_TYPE'] == project_type ]
    if df_info.shape[0] == 0 : init("Test type error: no sample ID corresponds.")

    df_info = df_info[~df_info['SAMPLE_ID'].str.contains('_PCE_|_NCE_|_PCT_|_NCT_', regex=True, na=False)]
    if df_info.shape[0] == 0 : init("No clinical specimens match the criteria.")

    uniq_info = fcDir_table(df_info, directory)
    if uniq_info.shape[0] == 0: init()

    df_info = pd.merge(df_info, uniq_info, on=['sub_name','PRJ_TYPE'])
    os.makedirs(os.path.join(outdir, df_info['seqDir'][0]), exist_ok=True)

    for pj_type in df_info['PRJ_TYPE'].unique():

        df_prj = df_info[df_info['PRJ_TYPE']==pj_type].reset_index()
        anal_dir = os.path.join(directory,pj_type,df_prj['seqDir'][0])


        if pj_type == 'eWES':
            out_file_1 = os.path.join(outdir, df_prj['seqDir'][0], pj_type + '.summarized.snv.target.xlsx')
            out_file_2 = os.path.join(outdir, df_prj['seqDir'][0], pj_type + '.target.snv.marked.xlsx')
            out_file_3 = os.path.join(outdir, df_prj['seqDir'][0], pj_type + '.cnv.marked.xlsx')
            remove_files([out_file_1,out_file_2,out_file_3])
        elif pj_type == 'WTS':
            out_file_1 = os.path.join(outdir, df_prj['seqDir'][0], pj_type + '.fusion.marked.xlsx')
            out_file_2 = os.path.join(outdir, df_prj['seqDir'][0], pj_type + '.cis-sage.filtered.xlsx')
            out_file_3 = os.path.join(outdir, df_prj['seqDir'][0], pj_type + '.exon_skipped.xlsx')
            remove_files([out_file_1,out_file_2,out_file_3])

        else :
            print('PRJ_TYPE error:' + pj_type )
            continue

        for i, item in df_prj.iterrows() :
#            if item['SAMPLE_ID'] == 'CR_25_00315_FP_R_JFN_1' : continue
            if pj_type == 'eWES':
                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID']+'.summarized.snv.target.tsv')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                    continue
                data = pd.read_csv(file, sep="\t")
                data = data[use_columns_ewes].drop_duplicates()
                data['Clinvar_CLNSIG'] = data['Clinvar_CLNSIG'].astype(str).replace("nan", None)
                data['ONCOKB_ONCOGENICITY'] = data['ONCOKB_ONCOGENICITY'].astype(str).replace("nan", None)
                data['REPORT'] = np.where(data['Clinvar_CLNSIG'].str.contains('Pathogenic|Likely_pathogenic', case=True, na=False, regex=False),'PASS',
                                np.where(data['ONCOKB_ONCOGENICITY'].str.contains('oncogenic|resistance', case=False, na=False, regex=False),'PASS',''))
                data = pd.merge(data, var_tbl, how='left', on='SYMBOL')
#                data['ONCOKB_REFSEQ_MATCH'] = np.where((data['ONCOKB_REFSEQ'].notna()) & (data['MANE_SELECT'].str.split('.').str[0] != data['ONCOKB_REFSEQ']),'MISMATCH', '')
#                data['ONCOKB_GENCODE_MATCH'] = np.where((data['ONCOKB_GENCODE'].notna()) & (data['Feature'].notna()) & (data['Feature'].str.split('.').str[0] != data['ONCOKB_GENCODE']),'MISMATCH', '')

                try:
                    with pd.ExcelWriter(out_file_1, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                except FileNotFoundError:
                    with pd.ExcelWriter(out_file_1, engine='openpyxl') as writer:
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                TMB_status = pic_value(os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID']+'.summarized.tmb.exome.tsv'), 'TMB_STATUS')
                MSI_status = pic_value(os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID']+'.summarized.msi.exome.tsv'), 'MSI_STATUS')
                MSI_status = 'MSI-H not detected' if MSI_status == 'MSS' else MSI_status
                TMB_score = pic_value(os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID']+'.summarized.tmb.exome.tsv'), 'TMB')
                MSI_score = pic_value(os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID']+'.summarized.msi.exome.tsv'), 'MSI')
                SAMPLING_DATE = '' if pd.isna(item['SAMPLING_DATE']) else item['SAMPLING_DATE'].strftime('%Y/%m/%d')

                coloring(out_file_1, item['SAMPLE_ID'], 'REPORT', 'PASS', "FFFF00")
                add_filt(out_file_1, item['SAMPLE_ID'])
#                add_info(out_file_1, item['SAMPLE_ID'], [item['PATH_NO'],item['DIAGNOSIS_NAME'],item['SAMPLING_DATE'].strftime('%Y/%m/%d'), [TMB_status,MSI_status], [TMB_score,MSI_score]])
                add_info(out_file_1, item['SAMPLE_ID'], [['Specimen_ID',item['PATH_NO']], ['Cancer_Type',item['DIAGNOSIS_NAME']], ['Biopsy/Surgery_Date',SAMPLING_DATE], [TMB_status,TMB_score],[MSI_status,MSI_score]])
                cull_columns(out_file_1, item['SAMPLE_ID'], hidden_columns_ewes_1, data)

                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'SNV', 'somatic', item['SAMPLE_ID']+'.target.snv.marked.tsv')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                    continue
                data = pd.read_csv(file, sep="\t")
                data = data[use_columns_ewes].drop_duplicates()
                try:
                    with pd.ExcelWriter(out_file_2, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                except FileNotFoundError:
                    with pd.ExcelWriter(out_file_2, engine='openpyxl') as writer:
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)

                add_filt(out_file_2, item['SAMPLE_ID'])
                cull_columns(out_file_2, item['SAMPLE_ID'], hidden_columns_ewes_2, data)

                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'CNV', item['SAMPLE_ID']+'.cnv.marked.tsv')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                    continue
                data = pd.read_csv(file, sep="\t", low_memory=False)
                data = data[ data['FILTER_GENES']=='PASS' ]
                data = data[use_columns_ewes_2].drop_duplicates()
                try:
                    with pd.ExcelWriter(out_file_3, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                except FileNotFoundError:
                    with pd.ExcelWriter(out_file_3, engine='openpyxl') as writer:
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)

                coloring(out_file_3, item['SAMPLE_ID'], ['FILTER_CN','FILTER_ONCOKB','FILTER_GENES'], 'PASS', "FFFF00")
                cull_columns(out_file_3, item['SAMPLE_ID'], hidden_columns_ewes_3, data)
                add_filt(out_file_3, item['SAMPLE_ID'])
                auto_filtering(out_file_3, item['SAMPLE_ID'], 'FILTER_CN', 'PASS')

            elif pj_type == 'WTS':

                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'Fusion', item['SAMPLE_ID']+'.fusion.marked.tsv')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                    data_f = pd.DataFrame(columns=use_columns_wts_1)
                else:
                    data = pd.read_csv(file, sep="\t")
                    data_f = data[['FILTER_ONCOKB','gene1','gene2','chr1','breakpoint_1','chr2','breakpoint_2','max_split_cnt','max_span_cnt','sample_type','disease','tools','inferred_fusion_type','samples','cancer_db_hits','fusion_IDs']]
                    data_f = data_f.drop_duplicates()
                    data_f = expand_breakpoints(data_f)
                    data_f = data_f.rename(columns={'FILTER_ONCOKB':'OncoKB'})
                    data_f.insert(1,'cancer-related',"")
                    data_f.insert(0,'Out-of-Frame',"PASS")
                    data_f['cancer_db_hits'] = data_f['cancer_db_hits'].astype(str)

                    try:
                        with pd.ExcelWriter(out_file_1, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                            data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                    except FileNotFoundError:
                        with pd.ExcelWriter(out_file_1, engine='openpyxl') as writer:
                            data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)

                    add_filt(out_file_1, item['SAMPLE_ID'])

                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'Fusion', 'Metafusion', 'final.n2.cluster.CANCER_FUSIONS.cis-sage.filtered')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                    data = pd.DataFrame()
                else :
                    data = pd.read_csv(file, sep="\t")
                    data = data.rename(columns={'#gene1':'gene1'})
                    data['cancer_db_hits'] = data['cancer_db_hits'].astype(str)
                    data = expand_breakpoints(data)

                if data.shape[0] > 0:
                    data = pd.merge(data, data_f, on=data.columns.tolist(), how='outer')[use_columns_wts_1]
                    data['Out-of-Frame'] = data['Out-of-Frame'].fillna('FAIL')
                    data = data.sort_values(['gene1','gene2','breakpoint_1','breakpoint_2'])
                else :
                    data = data_f

                try :
                    with pd.ExcelWriter(out_file_2, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                except FileNotFoundError:
                    with pd.ExcelWriter(out_file_2, engine='openpyxl') as writer:
                        data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)

                add_filt(out_file_2, item['SAMPLE_ID'])
                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID'] + '.summarized.splice.tsv')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                    as_result = pd.DataFrame()
                else:
                    as_result = pd.read_csv(os.path.join(anal_dir, item['SAMPLE_ID'], 'Summary', item['SAMPLE_ID'] + '.summarized.splice.tsv'), sep="\t")
                    as_result = as_result[ as_result['ONCOGENICITY'].notnull() ]
                
                if as_result.shape[0] > 0:
                    as_result = as_result['spliceName'].tolist()
                    add_info(out_file_2, item['SAMPLE_ID'], [['Specimen_ID',item['PATH_NO']], ['Cancer_Type',item['DIAGNOSIS_NAME']]]+as_result)
                else :
                    add_info(out_file_2, item['SAMPLE_ID'], [['Specimen_ID',item['PATH_NO']], ['Cancer_Type',item['DIAGNOSIS_NAME']]])

                file = os.path.join(anal_dir, item['SAMPLE_ID'], 'Alternative_splicing', 'ESDetector', item['SAMPLE_ID']+'.exon_skipped.tsv')
                if not os.path.isfile(file):
                    print('file not exists: ' + file)
                else:
                    data = pd.read_csv(file, sep="\t")
                    if data.shape[0] > 0 :
                        try:
                            with pd.ExcelWriter(out_file_3, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                                data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)
                        except FileNotFoundError:
                            with pd.ExcelWriter(out_file_3, engine='openpyxl') as writer:
                                data.to_excel(writer, sheet_name=item['SAMPLE_ID'], index=False)

            if i % 10 == 9 :
                print ('Completed processing of ' +  str(i+1) + ' specimens.')


