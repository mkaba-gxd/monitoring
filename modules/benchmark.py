import os
import sys
import argparse
import pymysql
import pandas as pd
import datetime
import warnings
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .func import *

file_benchmarks = os.path.join(os.path.dirname(__file__), 'benchmarks.idx')

def dataType_changes(file, sheet_name, colnames, datatype):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])

    for col in colnames:
        col_idx = df.columns.get_loc(col)
        col_letter = get_column_letter(col_idx)
        for row in range(2, df.shape[0] + 2):
            cell = ws[f'{col_letter}{row}']
            cell.number_format = datatype

    wb.save(file)

def run_benchmark(args):

    flowcellid = args.flowcellid
    directory = args.directory
    project_type = args.project_type
    outdir = args.outdir
    inclusion = [x.strip() for x in args.inclusion.split(',') if not x.strip() == '']
    exclusion = [x.strip() for x in args.exclusion.split(',') if not x.strip() == '']

    if len(inclusion) > 0 and len(exclusion) > 0 :
        init('ERROR: Inclusion and exclusion cannot be specified simultaneously.')

    bcm_tbl = pd.read_csv(file_benchmarks, sep="\t")

    df_info = getinfo(SelectData(flowcellid))
    if df_info.shape[0] == 0 : init("No matching data found.")

    if len(inclusion) > 0:
        print ("inclusion sample:" + "\n".join(inclusion))
        df_info = df_info[ df_info['SAMPLE_ID'].isin(inclusion)]
        if df_info.shape[0] == 0 : init("No corresponding sample IDs.")

    if len(exclusion) > 0:
        print ("exclusion sample:" + ",".join(exclusion))
        df_info = df_info[ ~df_info['SAMPLE_ID'].isin(exclusion)]
        if df_info.shape[0] == 0 : init("No corresponding sample IDs.")

    df_info['PRJ_TYPE'] = df_info['PRJ_TYPE'].str.replace('EWES',"eWES")
    if project_type == "both" :
        df_info = df_info[ df_info['PRJ_TYPE'].isin(['eWES','WTS']) ]
        bcm_tbl = bcm_tbl[ bcm_tbl['type'].isin(['eWES','WTS']) ]
    else :
        df_info = df_info[ df_info['PRJ_TYPE'] == project_type ]
        bcm_tbl = bcm_tbl[ bcm_tbl['type'] == project_type ]

    if df_info.shape[0] == 0 : init("Test type error: no sample ID corresponds.")
    if bcm_tbl.shape[0] == 0 : init("No index found for inspection type.")

    uniq_info = fcDir_table(df_info, directory)
    if uniq_info.shape[0] == 0: init("No matching data found.")

    df_info = pd.merge(df_info, uniq_info, on=['sub_name','PRJ_TYPE'])
    os.makedirs(outdir, exist_ok=True)

    for pj_type in df_info['PRJ_TYPE'].unique():

        df_prj = df_info[df_info['PRJ_TYPE']==pj_type].reset_index()
        bcm_prj = bcm_tbl[ bcm_tbl['type']==pj_type ][['class','name']]

        anal_dir = os.path.join(directory,pj_type,df_prj['seqDir'][0])
        out_file = os.path.join(outdir, df_prj['seqDir'][0] + '.xlsx')
#        if os.path.isfile(out_file) : os.remove(out_file)

        for i, item in df_prj.iterrows() :
            time_values = []

            for j, fac in bcm_prj.iterrows() :

                file_path = os.path.join(anal_dir, item['SAMPLE_ID'], 'Benchmark', fac['class'], '.'.join([item['SAMPLE_ID'],fac['name'],'tsv']))

                try :
                    df = pd.read_csv(file_path, sep="\t", header=0)
                    if 'h:m:s' in df.columns:
                        time_str = df['h:m:s'].iloc[0]
                        try :
                            parsed_time = datetime.datetime.strptime(time_str, '%H:%M:%S').time()
                        except Exception as e:
                            m, s = divmod(df['s'].iloc[0], 60)
                            h, m = divmod(m, 60)
                            parsed_time = str(int(h)) + ':' + str(int(m)) + ':' + str(int(s))
                    else :
                        parsed_time = 'unclear'

                except Exception as e:
                    parsed_time = '-'

                time_values.append(parsed_time)

            bcm_prj.insert(bcm_prj.shape[1], item['SAMPLE_ID'], time_values)

        try:
            with pd.ExcelWriter(out_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer :
                bcm_prj.to_excel(writer, sheet_name=pj_type, index=False)

        except FileNotFoundError:
            with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                bcm_prj.to_excel(writer, sheet_name=pj_type, index=False)

        dataType_changes(out_file, pj_type, df_prj['SAMPLE_ID'], 'hh:mm:ss')


